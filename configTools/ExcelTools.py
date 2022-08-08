#coding:utf-8
from openpyxl import load_workbook
from openpyxl import  Workbook
from configparser import ConfigParser

excleFilePath="D:\\case.xlsx"

wb=load_workbook(excleFilePath)

#获取所有工作表的名字
print ("所有的工作表的名字为：",wb.sheetnames)

# 创建指定工作表对象
# sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb['Login']

# 获取指定单元格的值
print (sheet['A1'].value)                 #获取第一行第二个单元格
print (sheet.cell(row=1,column=2).value)  #获取到第2行，第三列

#获取到已经使用的行数据
print ("sheet中已经使用的行数:",sheet.max_row)

#获取sheet中已经使用到列数
print ("sheet中已经使用到的列数为：",sheet.max_column)


#获取工作表的名字
print  ("工作表的名字为：",sheet.title)


#修改当前工作表名字
sheet.title="newSheet"


#修改指定的单元格的值
#sheet['C2'].value="新的值"
#sheet.cell(row=1,column=1,value='testtset')

#保存工作簿
wb.save(excleFilePath)

#关闭工作簿名称
wb.close()
